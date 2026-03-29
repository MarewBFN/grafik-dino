import calendar
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def _format_hour(time_str):
    if not time_str:
        return ""
    if str(time_str).endswith(":00"):
        return str(int(str(time_str).split(":")[0]))
    return str(time_str)

def export_schedule_to_excel(schedule, year, month, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Grafik"

    days_in_month = calendar.monthrange(year, month)[1]
    weekdays = ["Pn", "Wt", "Śr", "Cz", "Pt", "S", "N"]

    # Style
    fill_sat = PatternFill(start_color="E1E1E1", end_color="E1E1E1", fill_type="solid")
    fill_sun = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    sum_col_start = days_in_month + 3

    # NAGŁÓWEK GŁÓWNY (odwzorowanie ImageExporter)
    ws.cell(row=1, column=1, value=f"Grafik planowany {month:02d}/{year}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="Komórka:").font = Font(bold=True, size=12)
    
    ws.cell(row=1, column=sum_col_start-3, value="Wydruk wewnętrzny").font = Font(bold=True)
    ws.cell(row=2, column=sum_col_start-3, value=f"Data: {datetime.now().strftime('%d/%m/%Y')}").font = Font(bold=True)
    ws.cell(row=3, column=sum_col_start-3, value="Wygenerowany w Dingo")

    # NAGŁÓWEK TABELI
    ws.merge_cells("A4:B5")
    ws.cell(row=4, column=1, value="Nazwisko i imię").alignment = align_center
    
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=sum_col_start-1)
    ws.cell(row=3, column=3, value="Dni miesiąca").alignment = align_center
    ws.cell(row=3, column=3).font = Font(bold=True)
    
    # Nagłówki dni
    for day in range(1, days_in_month + 1):
        col = day + 2
        wd = calendar.weekday(year, month, day)
        ws.cell(row=4, column=col, value=day).alignment = align_center
        ws.cell(row=5, column=col, value=weekdays[wd]).alignment = align_center
        if wd == 5:
            ws.cell(row=4, column=col).fill = fill_sat
            ws.cell(row=5, column=col).fill = fill_sat
        elif wd == 6:
            ws.cell(row=4, column=col).fill = fill_sun
            ws.cell(row=5, column=col).fill = fill_sun

    # NAGŁÓWKI PODSUMOWANIA (odwzorowanie ImageExporter)
    sum_labels = ["Godziny", "Urlop", "L4", "Razem"]
    for i, lbl in enumerate(sum_labels):
        col = sum_col_start + i
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        c = ws.cell(row=4, column=col, value=lbl)
        c.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")
        c.font = Font(bold=True)

    # DANE PRACOWNIKÓW
    cur_row = 6
    for emp in schedule.employees:
        # Scalone nazwisko
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row+2, end_column=1)
        name_cell = ws.cell(row=cur_row, column=1, value=emp.display_name())
        name_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        name_cell.font = Font(bold=True)

        for i, lbl in enumerate(["od", "do", "h"]):
            ws.cell(row=cur_row + i, column=2, value=lbl).alignment = align_center

        # Dni miesiąca
        for day in range(1, days_in_month + 1):
            col = day + 2
            ds = schedule.get_day(emp, day)
            wd = calendar.weekday(year, month, day)
            
            cells = [ws.cell(row=cur_row + i, column=col) for i in range(3)]
            for c in cells:
                c.alignment = align_center
                if wd == 5:
                    c.fill = fill_sat
                elif wd == 6:
                    c.fill = fill_sun

            if ds.is_leave:
                ws.merge_cells(start_row=cur_row, start_column=col, end_row=cur_row+2, end_column=col)
                cells[0].value = "URL"
                cells[0].font = Font(bold=True)
            elif getattr(ds, "is_sick", False):
                ws.merge_cells(start_row=cur_row, start_column=col, end_row=cur_row+2, end_column=col)
                cells[0].value = "L4"
                cells[0].font = Font(bold=True)
            elif not ds.is_empty():
                cells[0].value = _format_hour(ds.start)
                cells[1].value = _format_hour(ds.end)
                cells[2].value = ds.total_as_str()

        # PODSUMOWANIE W WIERSZU (na prawo)
        values = [
            schedule.total_hours_for_employee(emp),
            schedule.leave_hours_for_employee(emp),
            schedule.sick_hours_for_employee(emp),
            schedule.total_with_leave_and_sick_for_employee(emp)
        ]

        for i, val in enumerate(values):
            col = sum_col_start + i
            ws.merge_cells(start_row=cur_row, start_column=col, end_row=cur_row+2, end_column=col)
            ws.cell(row=cur_row, column=col, value=val).alignment = align_center

        cur_row += 3

    # Krawędzie i szerokości
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 5
    for col_idx in range(1, sum_col_start + 4):
        col_ltr = get_column_letter(col_idx)
        if col_idx > 2: ws.column_dimensions[col_ltr].width = 6
        # Krawędzie tylko dla tabeli (od wiersza 3), omijamy nagłówek główny
        for r_idx in range(3, cur_row):
            # Omijamy puste pola koło "Dni miesiąca"
            if r_idx == 3 and (col_idx < 3 or col_idx >= sum_col_start):
                continue
            ws.cell(row=r_idx, column=col_idx).border = border_thin

    wb.save(path)